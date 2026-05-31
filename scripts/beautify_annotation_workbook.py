"""Beautify salesperson annotation workbooks without changing their schema.

The script keeps the original sheet names, header names, cell values, data
validations, and embedded images. It only adjusts visual presentation so the
workbook is easier for human annotators to read.
"""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path


GROUP_FILLS = {
    "material": "DCEBFF",
    "annotator": "E5E7EB",
    "state": "DFF7EA",
    "intervention": "FFF0C2",
    "speech": "EDE7FF",
    "score": "FFE4D6",
    "quality": "E7EEF8",
}

REQUIRED_INPUT_FILL = "FFFBEA"
PREFILLED_FILL = "F8FAFC"
EVEN_ROW_FILL = "FCFCFD"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("source", type=Path, help="Existing annotation workbook.")
    parser.add_argument("output", type=Path, help="Beautified workbook path.")
    parser.add_argument("--image-width", type=int, default=250)
    parser.add_argument("--image-height", type=int, default=141)
    return parser.parse_args()


def header_group(index: int) -> str:
    if 1 <= index <= 5:
        return "material"
    if 6 <= index <= 8:
        return "annotator"
    if 9 <= index <= 13:
        return "state"
    if 14 <= index <= 18:
        return "intervention"
    if 19 <= index <= 27:
        return "speech"
    if 28 <= index <= 37:
        return "score"
    return "quality"


def column_width(header: str) -> float:
    if header == "样本ID":
        return 17
    if header in {"第1张图", "第2张图", "第3张图"}:
        return 36
    if header == "场景描述":
        return 48
    if header in {"标注员ID", "销售经验年限", "标注日期"}:
        return 14
    if "适合度" in header or "确定度" in header or "紧迫程度" in header or "难不难" in header:
        return 13
    if header in {"你认为最合适的主动作", "你判断顾客所处阶段", "现在是否应该主动介入"}:
        return 20
    if "原因" in header or "反应" in header or "信号" in header or "备注" in header or "质量问题" in header:
        return 34
    if header.startswith("话术片段") or header.endswith("对应目的"):
        return 24
    return 24


def is_required_input_header(header: str) -> bool:
    return header not in {"样本ID", "第1张图", "第2张图", "第3张图", "场景描述", "标注员ID"}


def beautify(source: Path, output: Path, *, image_width: int, image_height: int) -> None:
    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ModuleNotFoundError as exc:
        raise SystemExit("openpyxl is required: uv run --with openpyxl python scripts/beautify_annotation_workbook.py ...") from exc

    output.parent.mkdir(parents=True, exist_ok=True)
    if source.resolve() != output.resolve():
        shutil.copy2(source, output)

    wb = load_workbook(output)
    thin = Side(style="thin", color="CBD5E1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    font_name = "PingFang SC"

    if "填写说明" in wb.sheetnames:
        ws = wb["填写说明"]
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        for col in range(1, min(ws.max_column, 4) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 24 if col == 1 else 58
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
                cell.font = Font(name=font_name, size=11, color="111827", bold=cell.row == 1 or cell.column == 1)
                cell.border = border
                if cell.row == 1:
                    cell.fill = PatternFill("solid", fgColor="1D4ED8")
                    cell.font = Font(name=font_name, size=16, color="FFFFFF", bold=True)
                elif cell.column == 1:
                    cell.fill = PatternFill("solid", fgColor="E8F1FF")
        for row_idx in range(1, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 30 if row_idx > 1 else 40

    for ws in wb.worksheets:
        if not ws.title.startswith("售货员"):
            continue
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "F2"
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.sheet_properties.outlinePr.summaryRight = False

        for idx, cell in enumerate(ws[1], start=1):
            fill = GROUP_FILLS[header_group(idx)]
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.font = Font(name=font_name, bold=True, color="111827", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            if cell.value:
                if "1-5" in str(cell.value):
                    cell.comment = Comment("请填写 1-5 分；5=最适合/最确定，1=最不适合/最不确定。", "PIWM")
                elif cell.value in {"你认为最合适的主动作", "现在是否应该主动介入", "你判断顾客所处阶段"}:
                    cell.comment = Comment("请从下拉选项中选择，不要手动输入新类别。", "PIWM")

        ws.row_dimensions[1].height = 44
        for idx, cell in enumerate(ws[1], start=1):
            ws.column_dimensions[get_column_letter(idx)].width = column_width(str(cell.value or ""))

        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = max(122, image_height * 0.76)
            row_fill = PatternFill("solid", fgColor=EVEN_ROW_FILL if row_idx % 2 == 0 else "FFFFFF")
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row_idx, col_idx)
                header = str(ws.cell(1, col_idx).value or "")
                cell.font = Font(name=font_name, color="111827", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border
                if col_idx <= 5:
                    cell.fill = row_fill
                elif is_required_input_header(header):
                    cell.fill = PatternFill("solid", fgColor=REQUIRED_INPUT_FILL)
                else:
                    cell.fill = PatternFill("solid", fgColor=PREFILLED_FILL)

        # Keep the workbook schema unchanged, but make embedded frames large enough
        # for annotators to judge gaze, posture, and facial expression.
        for image in getattr(ws, "_images", []):
            image.width = image_width
            image.height = image_height
            if hasattr(image.anchor, "ext"):
                image.anchor.ext.cx = image_width * 9525
                image.anchor.ext.cy = image_height * 9525

        # Group long text blocks visually; they remain expanded by default.
        for start, end in [("I", "M"), ("N", "R"), ("S", "AA"), ("AB", "AK"), ("AL", "AS")]:
            ws.column_dimensions.group(start, end, hidden=False)

        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output)


def main() -> None:
    args = parse_args()
    beautify(args.source, args.output, image_width=args.image_width, image_height=args.image_height)


if __name__ == "__main__":
    main()
