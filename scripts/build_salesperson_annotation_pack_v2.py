"""Build the Chinese salesperson annotation pack for balanced target 5-act test.

This pack is designed for Excel / Feishu import. It intentionally keeps PIWM
system labels out of the annotator-facing template and writes them to a
separate holdout file for later analysis.
"""

from __future__ import annotations

import argparse
import csv
import json
import shutil
import sys
from pathlib import Path
from typing import Any


ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from piwm_data.annotation_act_terms import CN_ACT_OPTIONS, EN_TO_CN_ACT, OPERATIONAL_5ACTS

DEFAULT_ACTION_EVAL = ROOT / "data/official/domain_specialization_eval_v2/target_frontcam_5act_test_action_selection.jsonl"
DEFAULT_MAIN_SCHEMA = ROOT / "data/official/piwm_target_v1/main_schema.jsonl"
DEFAULT_OUT_DIR = ROOT / "data/official/annotation_pack_v2"
FIVE_ACTS = list(OPERATIONAL_5ACTS)
ACT_ZH = EN_TO_CN_ACT
ACT_DROPDOWN_OPTIONS = list(CN_ACT_OPTIONS)


ANNOTATOR_COLUMNS = [
    "样本ID",
    "第1张图",
    "第2张图",
    "第3张图",
    "场景描述",
    "标注员ID",
    "销售经验年限",
    "标注日期",
    "你看到的顾客信号",
    "你判断顾客现在的状态",
    "你判断顾客所处阶段",
    "阶段判断确定度(1-5)",
    "阶段判断补充说明",
    "现在是否应该主动介入",
    "为什么现在应该/不应该介入",
    "介入紧迫程度(1-5)",
    "你认为最合适的主动作",
    "主动作判断确定度(1-5)",
    "如果介入，你第一句话会怎么说",
    "这句话的主要目的",
    "这句话还顺带包含哪些目的",
    "话术片段1",
    "片段1对应目的",
    "话术片段2",
    "片段2对应目的",
    "话术片段3",
    "片段3对应目的",
    f"{ACT_ZH['Greet']}适合度(1-5)",
    f"{ACT_ZH['Greet']}不适合原因(≤2时填写)",
    f"{ACT_ZH['Elicit']}适合度(1-5)",
    f"{ACT_ZH['Elicit']}不适合原因(≤2时填写)",
    f"{ACT_ZH['Inform']}适合度(1-5)",
    f"{ACT_ZH['Inform']}不适合原因(≤2时填写)",
    f"{ACT_ZH['Recommend']}适合度(1-5)",
    f"{ACT_ZH['Recommend']}不适合原因(≤2时填写)",
    f"{ACT_ZH['Hold']}适合度(1-5)",
    f"{ACT_ZH['Hold']}不适合原因(≤2时填写)",
    "为什么不选择第二合适做法",
    "你预计顾客接下来会怎么反应",
    "你认为的正向信号",
    "你认为的负向信号",
    "这条难不难标(1-5)",
    "图片/视频有没有质量问题",
    "是否需要讨论",
    "备注",
]


def read_jsonl(path: Path) -> list[dict[str, Any]]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]


def write_csv(path: Path, rows: list[dict[str, Any]], columns: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({column: row.get(column, "") for column in columns})


def write_xlsx(
    path: Path,
    rows: list[dict[str, Any]],
    columns: list[str],
    *,
    embed_images: bool = False,
    image_base_dir: Path | None = None,
) -> bool:
    """Write the three-annotator Excel workbook used for Feishu import."""
    try:
        from openpyxl import Workbook
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ModuleNotFoundError:
        return False

    workbook = Workbook()
    guide = workbook.active
    guide.title = "填写说明"
    guide_rows = [
        ["PIWM 售货员中文标注包 v2", "", "", ""],
        ["用途", "让 3 位售货员独立判断 30 条智能售货机前顾客片段是否应该主动介入，以及如何介入。", "", ""],
        ["主动作", "；".join(f"{ACT_ZH[act]}={act}" for act in FIVE_ACTS) + "。", "", ""],
        ["重要规则", "售货员不要看系统理论标签，只根据三张图和自己的经验填写。", "", ""],
        ["适合度", "5 个动作都要打 1-5 分；分数 ≤2 时填写不适合原因。", "", ""],
        ["话术拆分", "把第一句话尽量拆成 1-3 个片段，并给每个片段选择对应目的。", "", ""],
        ["反事实", "“为什么不选择第二合适做法”用于记录专家为什么没有选次优做法。", "", ""],
        ["导入飞书", "可直接导入本 xlsx；三位售货员分别填写各自 sheet。", "", ""],
        ["注意", "不要导入 theory_labels_holdout.csv 给售货员。", "", ""],
        ["阶段选项", "刚注意到商品/设备；开始感兴趣/浏览比较；已经有明确倾向；准备购买/确认", "", ""],
        ["介入选项", "应该介入；不应该介入；不确定", "", ""],
        ["生成来源", "data/official/domain_specialization_eval_v2/target_frontcam_5act_test_action_selection.jsonl", "", ""],
        ["样本数", "30 条，每条 3 张图，共 90 张图。", "", ""],
    ]
    for row in guide_rows:
        guide.append(row)
    guide["A1"].font = Font(bold=True, size=14)
    guide["A1"].fill = PatternFill("solid", fgColor="E8F1FF")
    guide.column_dimensions["A"].width = 18
    for col in ["B", "C", "D"]:
        guide.column_dimensions[col].width = 38
    for row in guide.iter_rows(min_row=1, max_row=len(guide_rows), min_col=1, max_col=4):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.font = Font(name="Arial", bold=cell.font.bold, size=cell.font.sz or 11)

    header_fill = PatternFill("solid", fgColor="E8F1FF")
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def col_letter(header: str) -> str:
        return get_column_letter(columns.index(header) + 1)

    def add_list_validation(sheet, header: str, values: list[str]) -> None:
        col = col_letter(header)
        validation = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
        sheet.add_data_validation(validation)
        validation.add(f"{col}2:{col}{len(rows) + 1}")

    def add_score_validation(sheet, header: str) -> None:
        col = col_letter(header)
        validation = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=True)
        validation.error = "该字段只能填写 1 到 5 的整数。"
        validation.errorTitle = "请输入 1-5"
        sheet.add_data_validation(validation)
        validation.add(f"{col}2:{col}{len(rows) + 1}")

    for sheet_name, annotator_id in [("售货员A", "S001"), ("售货员B", "S002"), ("售货员C", "S003")]:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(columns)
        for row in rows:
            sheet.append([
                annotator_id if column == "标注员ID" else row.get(column, "")
                for column in columns
            ])

        sheet.freeze_panes = "F2"
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = Font(name="Arial", bold=True, color="111827", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in sheet.iter_rows(min_row=2, max_row=len(rows) + 1, min_col=1, max_col=len(columns)):
            for cell in row:
                cell.font = Font(name="Arial", color="111827", size=10)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border

        for idx, header in enumerate(columns, 1):
            letter = get_column_letter(idx)
            if header in {"第1张图", "第2张图", "第3张图"} and embed_images:
                sheet.column_dimensions[letter].width = 24
            elif header in {"样本ID", "第1张图", "第2张图", "第3张图", "场景描述"}:
                sheet.column_dimensions[letter].width = 24
            elif "适合度" in header or "确定度" in header or "紧迫程度" in header or "难不难" in header:
                sheet.column_dimensions[letter].width = 17
            elif "原因" in header or "反应" in header or "信号" in header or "备注" in header or "质量问题" in header:
                sheet.column_dimensions[letter].width = 30
            else:
                sheet.column_dimensions[letter].width = 22

        add_list_validation(sheet, "你判断顾客所处阶段", ["刚注意到商品/设备", "开始感兴趣/浏览比较", "已经有明确倾向", "准备购买/确认"])
        add_list_validation(sheet, "现在是否应该主动介入", ["应该介入", "不应该介入", "不确定"])
        add_list_validation(sheet, "你认为最合适的主动作", ACT_DROPDOWN_OPTIONS)
        add_list_validation(sheet, "这句话的主要目的", ["破冰建立关系", "了解顾客需求", "传递产品信息", "推动具体决策", "缓解顾客顾虑", "其他"])
        for header in ["片段1对应目的", "片段2对应目的", "片段3对应目的"]:
            add_list_validation(sheet, header, ACT_DROPDOWN_OPTIONS)
        add_list_validation(sheet, "是否需要讨论", ["否", "是"])
        for header in [
            "阶段判断确定度(1-5)",
            "介入紧迫程度(1-5)",
            "主动作判断确定度(1-5)",
            f"{ACT_ZH['Greet']}适合度(1-5)",
            f"{ACT_ZH['Elicit']}适合度(1-5)",
            f"{ACT_ZH['Inform']}适合度(1-5)",
            f"{ACT_ZH['Recommend']}适合度(1-5)",
            f"{ACT_ZH['Hold']}适合度(1-5)",
            "这条难不难标(1-5)",
        ]:
            add_score_validation(sheet, header)

        if embed_images and image_base_dir is not None:
            for row_index, row in enumerate(rows, start=2):
                sheet.row_dimensions[row_index].height = 82
                for image_header in ("第1张图", "第2张图", "第3张图"):
                    relative_path = row.get(image_header)
                    if not relative_path:
                        continue
                    image_path = image_base_dir / relative_path
                    if not image_path.exists():
                        continue
                    image = ExcelImage(str(image_path))
                    image.width = 128
                    image.height = 72
                    sheet.add_image(image, f"{col_letter(image_header)}{row_index}")

    workbook.save(path)
    return True


def load_main_records(path: Path) -> dict[str, dict[str, Any]]:
    return {row["state_id"]: row for row in read_jsonl(path)}


def build_scene_description(main: dict[str, Any], *, scene_prefix: str) -> str:
    visual = main.get("visual_state") or {}
    summary = (visual.get("summary") or "").strip()
    if summary:
        return f"{scene_prefix}。{summary}"
    return scene_prefix


def build_rows(action_eval: Path, main_schema: Path, out_dir: Path) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    action_rows = read_jsonl(action_eval)
    main_by_id = load_main_records(main_schema)
    frames_dir = out_dir / "frames"
    frames_dir.mkdir(parents=True, exist_ok=True)
    annotator_rows: list[dict[str, str]] = []
    holdout_rows: list[dict[str, Any]] = []

    seen: set[str] = set()
    for row in action_rows:
        sample_id = row["source_id"]
        if sample_id in seen:
            raise ValueError(f"duplicate sample_id in action eval: {sample_id}")
        seen.add(sample_id)
        main = main_by_id.get(sample_id)
        if not main:
            raise ValueError(f"missing main_schema record for {sample_id}")

        copied_frames: list[str] = []
        for idx, image_path in enumerate(row["images"]):
            src = Path(image_path)
            if not src.exists():
                raise FileNotFoundError(src)
            dst = frames_dir / f"{sample_id}_{idx}.jpg"
            shutil.copy2(src, dst)
            copied_frames.append(f"frames/{dst.name}")

        visual = main.get("visual_state", {})
        candidate_acts = sorted(set(row.get("meta", {}).get("candidate_action_acts", {}).values()))
        best_act = row.get("meta", {}).get("best_act") or main.get("best_action_spec", {}).get("act")
        if best_act not in FIVE_ACTS:
            raise ValueError(f"{sample_id} best_act is not in 5-act set: {best_act}")

        annotator_rows.append(
            {
                "样本ID": sample_id,
                "第1张图": copied_frames[0],
                "第2张图": copied_frames[1],
                "第3张图": copied_frames[2],
                "场景描述": build_scene_description(main, scene_prefix="顾客在智能售货机前"),
            }
        )
        holdout_rows.append(
            {
                "sample_id": sample_id,
                "dataset_split": "target_5act_test",
                "source_dataset": "PIWM-Target-Frontcam-v1",
                "source_sampling_policy": "balanced_5act_test_split_6_per_act",
                "annotation_analysis_role": "target_primary_eval",
                "kappa_report_groups": "target_5act_all;target_4act_without_greet",
                "system_stage_label": main.get("aida_stage"),
                "system_intent_label": main.get("intent"),
                "system_best_act": best_act,
                "system_best_act_zh": ACT_ZH[best_act],
                "theory_go_no_go": "不介入" if best_act == "Hold" else "介入",
                "candidate_acts": "|".join(candidate_acts),
                "candidate_acts_zh": "|".join(ACT_ZH[act] for act in candidate_acts if act in ACT_ZH),
                "visual_summary": visual.get("summary"),
                "engagement_pattern": visual.get("engagement_pattern"),
                "gaze_and_attention": visual.get("gaze_and_attention"),
                "body_and_hands": visual.get("body_and_hands"),
            }
        )

    return annotator_rows, holdout_rows


def write_readme(out_dir: Path) -> None:
    act_terms_slash = " / ".join(ACT_DROPDOWN_OPTIONS)
    act_mapping = "，".join(f"{ACT_ZH[act]}={act}" for act in FIVE_ACTS)
    readme = f"""# PIWM 售货员中文标注包 v2

本目录用于把 30 条 balanced target-frontcam 5-act test 样本导入 Excel 或飞书多维表格，让 3 位中国售货员独立标注。

## 当前口径

- 5 个主动作：{act_terms_slash}。
- 英文映射：{act_mapping}。
- 售货员只看图片和中文场景，不看系统 stage、系统 best act 或候选动作标签。

## 文件

- `annotation_template_single_annotator.csv`：单个售货员填写模板，30 行。
- `annotation_template_three_annotators.xlsx`：三张独立填写页，分别为 `售货员A`、`售货员B`、`售货员C`。
- `annotation_template_three_annotators_with_images.xlsx`：同样的三人填写模板，但已把 90 张图片直接嵌入表格，适合预览和人工分发表。
- `frames/`：90 张抽帧图，命名为 `{{sample_id}}_{{0/1/2}}.jpg`。
- `theory_labels_holdout.csv`：系统理论标签，不给售货员看，只用于后续一致性和偏好分析。
- `theory_labels_holdout.jsonl`：同上，JSONL 版本，便于脚本读取。

## 统计口径

- target 包含 5 个动作：问候 / 询问需求 / 介绍信息 / 推荐商品 / 继续观察等待。
- 和 general probe 对比专家一致性时，报告两套 κ：`target_5act_all` 使用全部 30 条；`target_4act_without_greet` 排除 system best act 为 `Greet` 的 6 条，只和 general probe 的 4-act 口径直接对齐。
- `theory_labels_holdout.csv/jsonl` 只用于分析，不导入给售货员。

## 重新生成

如果只需要 CSV、帧图和隐藏标签：

```bash
python3 scripts/build_salesperson_annotation_pack_v2.py
```

如果要同时重建 Excel 工作簿，确保 Python 环境有 `openpyxl`；本机有 `uv` 时可直接运行：

```bash
uv run --with openpyxl python scripts/build_salesperson_annotation_pack_v2.py
```

## 飞书导入建议

1. 优先导入 `annotation_template_three_annotators.xlsx`。
2. 给三位售货员分别开放 `售货员A`、`售货员B`、`售货员C` 对应表或视图。
3. 三位售货员不要互相看对方结果，避免相互影响。
4. `第1张图`、`第2张图`、`第3张图` 可以保留路径，也可以在飞书中改成附件字段后上传 `frames/` 目录里的图片。

## 填写规则

- `现在是否应该主动介入`：填写“应该介入 / 不应该介入 / 不确定”。
- `你认为最合适的主动作`：从“{act_terms_slash}”中选择一个。
- `这句话的主要目的`：填写这句话在沟通中的目的，例如“破冰建立关系”或“了解顾客需求”，不要写系统英文标签。
- 5 个适合度分数都必填，范围 1-5。
- 如果某个动作适合度 ≤ 2，请填写对应“不适合原因”。
- `为什么不选择第二合适做法`：第二合适做法由 5 个动作分数自动推导，售货员只需要写为什么最终没选它。
- 话术拆分最多填 3 段。例如：
  - 片段1：您好，看您在看饮料；片段1对应目的：问候
  - 片段2：是想看口味还是价格？片段2对应目的：询问需求

## 后处理

售货员完成后，从飞书导出 CSV，然后运行：

```bash
python3 scripts/convert_salesperson_annotation_export.py \\
  --input exported_salesperson_a.csv \\
  --holdout data/official/annotation_pack_v2/theory_labels_holdout.jsonl \\
  --output data/official/annotation_pack_v2/converted_salesperson_a.jsonl
```
"""
    (out_dir / "README.md").write_text(readme, encoding="utf-8")


def run(args: argparse.Namespace) -> dict[str, Any]:
    out_dir = args.output_dir
    out_dir.mkdir(parents=True, exist_ok=True)
    annotator_rows, holdout_rows = build_rows(args.action_eval, args.main_schema, out_dir)
    write_csv(out_dir / "annotation_template_single_annotator.csv", annotator_rows, ANNOTATOR_COLUMNS)
    xlsx_generated = write_xlsx(out_dir / "annotation_template_three_annotators.xlsx", annotator_rows, ANNOTATOR_COLUMNS)
    xlsx_with_images_generated = write_xlsx(
        out_dir / "annotation_template_three_annotators_with_images.xlsx",
        annotator_rows,
        ANNOTATOR_COLUMNS,
        embed_images=True,
        image_base_dir=out_dir,
    )
    write_csv(
        out_dir / "theory_labels_holdout.csv",
        holdout_rows,
        [
            "sample_id",
            "dataset_split",
            "source_dataset",
            "source_sampling_policy",
            "annotation_analysis_role",
            "kappa_report_groups",
            "system_stage_label",
            "system_intent_label",
            "system_best_act",
            "system_best_act_zh",
            "theory_go_no_go",
            "candidate_acts",
            "candidate_acts_zh",
            "visual_summary",
            "engagement_pattern",
            "gaze_and_attention",
            "body_and_hands",
        ],
    )
    with (out_dir / "theory_labels_holdout.jsonl").open("w", encoding="utf-8") as handle:
        for row in holdout_rows:
            handle.write(json.dumps(row, ensure_ascii=False, sort_keys=True) + "\n")
    write_readme(out_dir)
    return {
        "output_dir": str(out_dir),
        "samples": len(annotator_rows),
        "frames": len(list((out_dir / "frames").glob("*.jpg"))),
        "xlsx": str(out_dir / "annotation_template_three_annotators.xlsx"),
        "xlsx_generated": xlsx_generated,
        "xlsx_with_images": str(out_dir / "annotation_template_three_annotators_with_images.xlsx"),
        "xlsx_with_images_generated": xlsx_with_images_generated,
    }


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--action-eval", type=Path, default=DEFAULT_ACTION_EVAL)
    parser.add_argument("--main-schema", type=Path, default=DEFAULT_MAIN_SCHEMA)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUT_DIR)
    return parser


def main() -> None:
    print(json.dumps(run(build_parser().parse_args()), ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
